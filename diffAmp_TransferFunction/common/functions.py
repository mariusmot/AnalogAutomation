from datetime import datetime
import math
import os
import openpyxl
from openpyxl.styles import Font

class functions():  

    def get_variables_from_excel(self, file_path):       
        workbook = openpyxl.load_workbook(file_path)
        # Select the worksheet (by default, the first worksheet will be selected)
        worksheet = workbook.active
        dictionaries = []
        for col in worksheet.iter_cols(min_row=2, max_row=2):                
            cell_value = col[0].value                
            if cell_value:
                # Split the cell value using a delimiter (e.g., comma)
                pairs = cell_value.split(',')
                # Store the variables and their values in a dictionary
                variables = {}
                for pair in pairs:
                    variable_name, value = pair.strip().split(':')
                    variables[variable_name] = value
                dictionaries.append(variables)
        return dictionaries
    
    def copy_columns_between_excels(self,
        excel_file_1, excel_file_2,
        sheet_1, column1_1, column2_1,
        sheet_2, column1_2, column2_2):
        # Load the workbooks and worksheets
        wb1 = openpyxl.load_workbook(excel_file_1)
        ws1 = wb1[sheet_1]
        wb2 = openpyxl.load_workbook(excel_file_2)
        ws2 = wb2[sheet_2]
        # Copy the specified columns from sheet_1 to sheet_2
        for row in range(1, ws1.max_row + 1):
            ws2.cell(row=row, column=column1_2).value = ws1.cell(row=row, column=column1_1).value
            ws2.cell(row=row, column=column2_2).value = ws1.cell(row=row, column=column2_1).value
        # Save the changes to excel_file_2
        wb2.save(excel_file_2)
    
    def get_formatted_current_date(self):
        now = datetime.now()
        day = str(now.day)
        date = now.strftime(f"%B {day}, %Y")
        return date
    
    def value_to_position(self, value, limitmin, limitmax):
        minpos = 1
        maxpos = 10000
        minval = math.log(limitmin)
        maxval = math.log(limitmax)
        scale = (maxval - minval) / (maxpos - minpos)
        if value <= 0:
            return minpos
        else:
            position = minpos + (math.log(value) - minval) / scale
            return position
        
    def text_to_num(self, si_string):
        si_prefixes = {
            'y': 1e-24,  # yocto
            'z': 1e-21,  # zepto
            'a': 1e-18,  # atto
            'f': 1e-15,  # femto
            'p': 1e-12,  # pico
            'n': 1e-9,   # nano
            'u': 1e-6,   # micro
            'm': 1e-3,   # milli
            'k': 1e3,    # kilo
            'M': 1e6,    # mega
        }        
        if si_string[-1] in si_prefixes:
            value = float(si_string[:-1]) * si_prefixes[si_string[-1]]
        else:
            value = float(si_string)
        return value 
    
    def create_excel_file(self, folder_path, file_name):
        workbook = openpyxl.Workbook()
        file_path = os.path.join(folder_path, file_name)
        workbook.save(file_path)

    #This functions moves data from a sheet to another in the same excel file            
    def copy_ranges_within_excel(workbook, source_sheet, target_sheet, source_col1, source_col2, target_col1, target_col2, offset_source_sheet, offset_target_sheet):
        wb = openpyxl.load_workbook(workbook)
        ws_source = wb[source_sheet]
        ws_target = wb[target_sheet]
        max_row = ws_source.max_row

        for row in range(1 + offset_source_sheet, max_row + 1):
            # Copy data from source_col1
            cell_value = ws_source.cell(row=row, column=source_col1).value
            if cell_value is not None:
                ws_target.cell(row=row+offset_target_sheet-offset_source_sheet, column=target_col1).value = cell_value

            # Copy data from source_col2
            cell_value = ws_source.cell(row=row, column=source_col2).value
            if cell_value is not None:
                ws_target.cell(row=row+offset_target_sheet-offset_source_sheet, column=target_col2).value = cell_value
        wb.save(workbook)

        # # Call the function with appropriate arguments
        # copy_ranges_within_excel(workbook_path, 'Datasheet', gain_sheet_score, 5, 6, 3, 4, offset_source_sheet=1, offset_target_sheet=2)

    def get_min_and_max_range_values(self, file, sheet):
        wb2 = openpyxl.load_workbook(file)
        ws2 = wb2[sheet]
        max_row = ws2.max_row
        # Find the last non-empty value in column 5
        for i in range(max_row, 1, -1):
            cell_value_5 = ws2.cell(row=i, column=5).value
            if cell_value_5 is not None and cell_value_5 != '':
                last_value_col_5 = cell_value_5
                break
        else:
            last_value_col_5 = None
        # Find the last non-empty value in column 6
        for i in range(max_row, 1, -1):
            cell_value_6 = ws2.cell(row=i, column=6).value
            if cell_value_6 is not None and cell_value_6 != '':
                last_value_col_6 = cell_value_6
                break
        else:
            last_value_col_6 = None
        # Retrieve the first value in column 5
        first_value_col_5 = ws2.cell(row=2, column=5).value
        # Retrieve the first value in column 6
        first_value_col_6 = ws2.cell(row=2, column=6).value

        return first_value_col_5, last_value_col_5, first_value_col_6, last_value_col_6
    
    #This functions moves datasheet from a sheet to another in the same excel file            
    def copy_ranges_within_excel(self, workbook, source_sheet, target_sheet, source_col1, source_col2, target_col1, target_col2, offset_source_sheet, offset_target_sheet):
        wb = openpyxl.load_workbook(workbook)
        ws_source = wb[source_sheet]
        ws_target = wb[target_sheet]

        max_row = ws_source.max_row

        for row in range(1 + offset_source_sheet, max_row + 1):
            # Copy data from source_col1
            cell_value = ws_source.cell(row=row, column=source_col1).value
            if cell_value is not None:
                ws_target.cell(row=row+offset_target_sheet-offset_source_sheet, column=target_col1).value = cell_value

            # Copy data from source_col2
            cell_value = ws_source.cell(row=row, column=source_col2).value
            if cell_value is not None:
                ws_target.cell(row=row+offset_target_sheet-offset_source_sheet, column=target_col2).value = cell_value

        wb.save(workbook)

    #This fuction applies the formulas to create the score for Nimble and LTspice
    def apply_formulas(self, workbook, sheet1_name, sheet2_name, x_ax_min, x_ax_max, y_ax_min, y_ax_max):
        wb = openpyxl.load_workbook(workbook)

        sheet1 = wb[sheet1_name]
        sheet2 = wb[sheet2_name]

        max_row = sheet1.max_row
        column_C = 'C'

        # Find the last non-empty row in column G
        for row in range(max_row, 0, -1):
            if sheet1[f"{column_C}{row}"].value is not None:
                max_row = row
                break

        print(f"Max row in column G of sheet2: {max_row}")

        # Iterate through cells E3:E56 and F3:F56 in sheet1 and apply the formulas
        for row in range(3, max_row+1):
            cell_e = sheet1.cell(row=row, column=5)
            cell_e.value = f'=MATCH(C{row}, INDIRECT("\'{sheet2_name}\'!$A$2:$A$432"), 1)'
            cell_f = sheet1.cell(row=row, column=6)
            cell_f.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$A$2:$A$432"), E{row})'
            cell_g = sheet1.cell(row=row, column=7)
            cell_g.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$A$2:$A$432"), E{row}+1)'                
            cell_h = sheet1.cell(row=row, column=8)
            cell_h.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$B$2:$B$432"), E{row})'               
            cell_i = sheet1.cell(row=row, column=9)
            cell_i.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$B$2:$B$432"), E{row}+1)'              
            cell_j = sheet1.cell(row=row, column=10)
            cell_j.value = f'=SLOPE(H{row}:I{row}, F{row}:G{row})*(C{row}-F{row})+H{row}'              
            cell_k = sheet1.cell(row=row, column=11)
            cell_k.value = f'=ABS(J{row}-(D{row}))'

        for row in range(3, max_row+1):
            cell_m = sheet1.cell(row=row, column=13)  # Column 13 corresponds to 'M' =MATCH(C3,'G2'!$A$2:$A$432,1)
            cell_m.value = f'=MATCH(C{row}, INDIRECT("\'{sheet2_name}\'!$C$2:$C$1002"), 1)'
            cell_n = sheet1.cell(row=row, column=14)  # Column 14 corresponds to 'N' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3)
            cell_n.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$C$2:$C$1002"), M{row})'
            cell_o = sheet1.cell(row=row, column=15)  # Column 15 corresponds to 'O' =INDEX('G2'!$A$2:$A$432,'G2 Score'!E3+1)
            cell_o.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$C$2:$C$1002"), M{row}+1)'      
            cell_p = sheet1.cell(row=row, column=16)  # Column 16 corresponds to 'P' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3)
            cell_p.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$D$2:$D$1001"), M{row})'   
            cell_q = sheet1.cell(row=row, column=17)  # Column 17 corresponds to 'Q' =INDEX('G2'!$B$2:$B$432,'G2 Score'!E3+1)
            cell_q.value = f'=INDEX(INDIRECT("\'{sheet2_name}\'!$D$2:$D$1001"), M{row}+1)'   
            cell_r = sheet1.cell(row=row, column=18)  # Column 18 corresponds to 'R' =SLOPE(H3:I3,F3:G3)*(C3-F3)+H3
            cell_r.value = f'=SLOPE(P{row}:Q{row}, N{row}:O{row})*(C{row}-N{row})+P{row}'    
            cell_s = sheet1.cell(row=row, column=19)  # Column 19 corresponds to 'S' =ABS(J3-D3)
            cell_s.value = f'=ABS(R{row}-(D{row}))'
            
        #This function will determine valid data in range, on which scoring will be applied
        x_range = [float(x_ax_min), float(x_ax_max)] 
        y_range = [float(y_ax_min), float(y_ax_max)]     

        valid_rows = []
        for row in sheet1.iter_rows(min_row=3):
            if row[2].value is not None and row[3].value is not None and x_range[0] <= row[2].value <= x_range[1] and y_range[0] <= row[3].value <= y_range[1]:
                valid_rows.append(row[0].row)
                
        valid_rows_range_l = "K" + str(min(valid_rows)) + ":K" + str(max(valid_rows))
        print(valid_rows_range_l)

        valid_rows_range_t = "S" + str(min(valid_rows)) + ":S" + str(max(valid_rows))
        print(valid_rows_range_t)

        cell_l3 = sheet1.cell(row=3, column=12)  # Column 12 corresponds to 'L'
        cell_l3.value = f'=AVERAGE({valid_rows_range_l})'
        cell_l3.font = Font(bold=True)

        cell_t3 = sheet1.cell(row=3, column=20)  # Column 20 corresponds to 'T'
        cell_t3.value = f'=AVERAGE({valid_rows_range_t})'
        cell_t3.font = Font(bold=True)

        wb.save(workbook)
        