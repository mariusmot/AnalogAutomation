import unittest
import os
import shutil
import zipfile
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from decimal import Decimal
import math
import time
import json
import chromedriver_autoinstaller
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.chart.axis import ChartLines
import ltspice
from PyLTSpice import SimCommander
import numpy as np
from common.functions import functions

class diffAMP(unittest.TestCase):

    def setUp(self):
        # driver instance
        chromedriver_autoinstaller.install()
        options = Options()
        options.add_argument("--headless=new")
        self.driver = webdriver.Chrome(options=options) 

        
    def test_export(self):
        with open(r'DiffAmp_TransferFunction.json') as d:
            self.testData = json.load(d)['Nimble'][0]
        my_functions = functions() 
        device = self.testData['device']
        project_path = self.testData['project_location']
        source_workbook = (project_path + '\\' + device + '_WithScores.xlsx') 
        dictionaries = my_functions.get_variables_from_excel(source_workbook)
        results_folder = project_path + '\\' + 'Automated_Test_Results'
        results_file = results_folder + '\\' + device + "_Test_Results.xlsx"

        # Create results folder and results file
        if not os.path.exists(project_path + '\\' + 'Automated_Test_Results'):
            os.makedirs(project_path + '\\' + 'Automated_Test_Results')
        my_functions.create_excel_file(results_folder, results_file)

        for dictionary in dictionaries:
            chromedriver_autoinstaller.install()
            options = Options()
            options.add_argument("--headless=new")
            self.driver = webdriver.Chrome(options=options) 
        
            driver = self.driver
            driver.maximize_window()
            driver.get(self.testData['URL'])

            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("       DiffAmp_TransferFunction script is running        ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++") 

            device = self.testData['device']
            gain = dictionary['gain']
            R1 = dictionary['R1']
            C1 = dictionary['C1']
            R3 = dictionary['Load']
            filter_freq = dictionary['Filter_freq']
            downloads_path = self.testData['downloads_path']
            project_path = self.testData['project_location']
            
            source_workbook = project_path + '\\' + device + '_WithScores.xlsx'
             
            new_rvalue = my_functions.text_to_num(R1)
            new_c1_value = my_functions.text_to_num(C1) 
            # new_rc1value = my_functions.text_to_num(paths['rc1_value'])
            new_r3value = my_functions.text_to_num(R3) 
              
            # cookies accept
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#noise-spinner")))
            WebDriverWait(driver, 10).until(EC.invisibility_of_element((By.CSS_SELECTOR, "#noise-spinner")))
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "body.ember-application:nth-child(2) div.consent-dialog:nth-child(1) div.modal.fade.in.show "
                                "div.modal-dialog div.modal-content div.modal-body div.short-description > a.btn.btn-success:nth-child(2)"))).click()
            
            # amplifier settings
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((
                By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']/div[@id='build-signal-chain-tab-content']"
                "/div[@id='adi-signal-chain-row']/div[@id='analog-signal-chain-group']/div[@id='signal-chain-drop-area']/table[1]/tr[1]/td[1]/div[1]/div[2]/div[2]/*[1]"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).send_keys(Keys.CONTROL + "a")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).send_keys(Keys.DELETE)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).send_keys(gain)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#amp-gain-input"))).send_keys(Keys.ENTER)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#tspan2988-4-54-5"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-0"))).send_keys(device)

            #check if part is present or not disabled in Nimble list
            try:
                element = driver.find_element(By.CSS_SELECTOR, "#device-table > div.slick-pane.slick-pane-top.slick-pane-left > div.slick-viewport.slick-viewport-top.slick-viewport-left > div > div")
                class_attribute = element.get_attribute('class')
                if class_attribute and 'disabled' in class_attribute:
                    raise Exception(device + " can't be selected in Nimble list")
                else:
                    element.click()
            except NoSuchElementException:
                raise Exception(device + " can't be selected in Nimble list")
            
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, 'body.ember-application.modal-open:nth-child(2) div.adi-modal.modal-fills-window.modal-hide-scroll:nth-child(5) div.modal.fade.show.d-block:nth-child(1) '
                'div.modal-dialog div.modal-content div.modal-body div.configure-amp.configure-signal-chain-item div.adi-modal.modal-fills-window:nth-child(5) '
                'div.modal.fade.show.d-block:nth-child(1) div.modal-dialog div.modal-content div.modal-footer div.button-row > button.btn.btn-primary:nth-child(1)' ))).click()         
                     
            rposition = my_functions.value_to_position(new_rvalue, 1e1, 1e7)
            c1position = my_functions.value_to_position(new_c1_value, 1e-13, 1e-6)            
            # rc1position = value_to_position(new_rc1value, 150, 1e7)
            r3position = my_functions.value_to_position(new_r3value, 150, 1e7)

            driver.execute_script(f"document.querySelector('#rscale-slider').value = {rposition}; document.querySelector('#rscale-slider').dispatchEvent(new Event('input'));")
            driver.execute_script(f"document.querySelector('#c1-slider').value = {c1position}; document.querySelector('#c1-slider').dispatchEvent(new Event('input'));")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "#config-signal-chain-item-modal > div.modal.fade.show.d-block > div > div > div.modal-footer.signal-chain-base-modal-footer > div > button.btn.btn-primary"))).click()
            
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("                    Slider values set!                   ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")        

            #Set Up Filter values 
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#signal-chain-drop-area #circuit-content[title=\"Filter\"]"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-inputs-type-tab-button"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-type-category-radio-group > div:nth-child(3) > label > input[type=radio]"))).click()        
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#hp-diff-wiring-button"))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#filter-order-radio-group > div:nth-child(1) > label > input[type=radio]"))).click()        
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#fp-input"))).send_keys(Keys.CONTROL + "a")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#fp-input"))).send_keys(Keys.DELETE)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#fp-input"))).send_keys(filter_freq)
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "#config-signal-chain-item-modal > div.modal.fade.show.d-block > div > div > div.modal-body.signal-chain-base-modal-body > div > div.top-area"
                " > section.config-section > div > div.sub-tab-content-container > button.tab-button-area.enabled.next > div > svg"))).click()

            #driver.execute_script(f"document.querySelector('#rc-r1-slider').value = {rc1position}; document.querySelector('#rc-r1-slider').dispatchEvent(new Event('input'));")
            driver.execute_script(f"document.querySelector('#rc-r3-slider').value = {r3position}; document.querySelector('#rc-r3-slider').dispatchEvent(new Event('input'));")        
            
            driver.execute_script("document.querySelector('#config-signal-chain-item-modal > div.modal.fade.show.d-block > div > div > div.modal-footer > div > button.btn.btn-primary').click()")
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-tab"))).click()
            
            #This script is moving the downloaded files to the project folder
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-tab"))).click()
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-container > div.download-area > div.download-all-button > div > h5"))).click()
            time.sleep(5)         

            # today = date.today()
            # day = str(today.day)
            # current_date = today.strftime(f"%B {day}, %Y")
            current_date = my_functions.get_formatted_current_date()

            zip_file_path = downloads_path + '\\' + 'Full Data Export - ' + current_date + '.zip'
            with zipfile.ZipFile(zip_file_path) as zip_ref:
                new_path = project_path + '\\' + device + ' extracted files'
                zip_ref.extractall(new_path)
            print("Files were extracted to project folder")

            #Deletes the zip file after extracting
            if os.path.exists(zip_file_path):
                os.remove(zip_file_path)
            else:
                print("Zip file does not exist")   

            # Move Transfer Function csv to extracted files folder
            raw_data = project_path + '\\' + device + ' extracted files' + '\\' + 'Raw Data' + '\\' + 'Individual Stage Data' + '\\' + 'Amplifier' + '\\' + 'Amplifier - Transfer Function.csv'
            shutil.copy2(raw_data, new_path)

            # Converting the Transfer Function .csv to .xlsx
            path_file = pd.read_csv(project_path + '\\' + device + ' extracted files' + '\\' + 'Amplifier - Transfer Function.csv')
            path_file.to_excel(project_path + '\\' + device + ' extracted files' + '\\' + 'Amplifier - Transfer Function.xlsx', index=None, header=True)
            nimble_output_path = project_path + '\\' + device + ' extracted files' + '\\' + 'Amplifier - Transfer Function.xlsx'

            #Running the simulation in LTSpice
            file_path = project_path + '\\' + device + ' extracted files' + '\\' + 'Ltspice Schematics'
            LTC = SimCommander(file_path + "\\AC_Simulation.asc")
            LTC.run()
            LTC.wait_completion()
            
            # parsing LTSpice files
            l = ltspice.Ltspice(file_path + "\\AC_Simulation_1.raw")
            l.parse()

            # Get the V(out) trace data
            freq = l.get_frequency()
            Vout = l.get_data('V(out)')

            # Change from Carthesian to Polar format 
            Vout_dB = 20 * np.log10(np.abs(Vout))

            # Create a DataFrame with the frequency and magnitude (dB) data
            data = {'Frequency (Hz)': freq, 'Magnitude (dB)': Vout_dB}
            df = pd.DataFrame(data)

            # Export the DataFrame to an Excel file
            ltspice_output_path = (new_path + "\\AC_Simulation.xlsx")
            df.to_excel(ltspice_output_path, index=False, engine='openpyxl')    

            # Name the sheet
            file = openpyxl.load_workbook(results_file)
            sheet = file.active
            result_sheet = ("G" + gain + 'RL' + R3 + 'Ω' )
            sheet.title = result_sheet
            #sheet.delete_cols(3)
            file.save(results_file)

            # Getting the Nimble data from the Transfer_Function.xlsx to Result file
            my_functions.copy_columns_between_excels(
                nimble_output_path, results_file,
                'Sheet1', 1, 2,
                result_sheet, 1, 2) 

            # Getting the LTSpice data from the AC_Simulation.xlsx to Result file
            my_functions.copy_columns_between_excels(
                ltspice_output_path, results_file,
                'Sheet1', 1, 2,
                result_sheet, 3, 4)  
            
            # Getting the Datasheet data to Result File                   
            wb1 = openpyxl.load_workbook(source_workbook)
            ws1 = wb1['Datasheet']
            wb2 = openpyxl.load_workbook(results_file)
            ws2 = wb2[result_sheet]
            # Copy the specified columns from sheet_1 to sheet_2
            ws2.cell(row=3, column=5).value = None
            ws2.cell(row=3, column=6).value = None
            column1_1 = 1
            column2_1 = 2
            for row in range(1, ws1.max_row + 1):
                if row != 2:
                    ws2.cell(row=row, column=5).value = ws1.cell(row=row, column=column1_1).value
                    ws2.cell(row=row, column=6).value = ws1.cell(row=row, column=column2_1).value
            #Shift rows up
            for row in range(3, ws2.max_row + 1):
                ws2.cell(row=row-1, column=5).value = ws2.cell(row=row, column=5).value
                ws2.cell(row=row-1, column=6).value = ws2.cell(row=row, column=6).value
            # Clear values in the last row
            ws2.cell(row=ws2.max_row, column=5).value = None
            ws2.cell(row=ws2.max_row, column=6).value = None
            column1_1 += 2
            column2_1 += 2
            wb2.save(results_file)

            #get x and y ranges
            range_result = my_functions.get_min_and_max_range_values(results_file, result_sheet) 
            first_value_col_5, last_value_col_5, first_value_col_6, last_value_col_6 = range_result

            x_min = first_value_col_5
            y_min = min(first_value_col_6, last_value_col_6)
            x_max = last_value_col_5
            y_max = max(first_value_col_6, last_value_col_6)

            #Creating Scatter graph    
            workbook = load_workbook(results_file)
            sheet = workbook[result_sheet]
            link = driver.current_url
            sheet['J29'] = link

            sheet.cell(row=1, column=1).value = "Nimble-Freq."
            sheet.cell(row=1, column=2).value = "Nimble-Mag."
            sheet.cell(row=1, column=3).value = "LTSpice-Freq."
            sheet.cell(row=1, column=4).value = "LTSspice-Mag."
            sheet.cell(row=1, column=5).value = "Datasheet-Freq."
            sheet.cell(row=1, column=6).value = "Datasheet-Mag."

            for i in range(1,21):
                sheet.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True)

            x_nimble = Reference(sheet, min_col=2, min_row=2, max_row=1010)
            y_nimble = Reference(sheet, min_col=1, min_row=2, max_row=1010)
            x_ltspice = Reference(sheet, min_col=4, min_row=2, max_row=1010)
            y_ltspice = Reference(sheet, min_col=3, min_row=2, max_row=1010)
            x_datasheet = Reference(sheet, min_col=6, min_row=2, max_row=1010)
            y_datasheet = Reference(sheet, min_col=5, min_row=2, max_row=1010)

            series_nimble = Series(x_nimble, y_nimble, title_from_data=False, title="Nimble")
            series_ltspice = Series(x_ltspice, y_ltspice, title_from_data=False, title="LTspice")
            series_datasheet = Series(x_datasheet, y_datasheet, title_from_data=False, title="Datasheet")
            
            # Chart type
            chart = ScatterChart()
            chart.series.append(series_nimble)
            chart.series.append(series_ltspice)
            chart.series.append(series_datasheet)

            chart.x_axis.scaling.logBase = 10
            chart.y_axis.number_format = '0.00E+00'
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.tickLblSkip = 3

            # Calculate the range of your data
            x_data_range = last_value_col_5 - first_value_col_5
            print(x_data_range)
            y_data_range = last_value_col_6 - first_value_col_6
            print(y_data_range)

            chart.x_axis.scaling.min = x_min
            chart.y_axis.scaling.min = y_min
            chart.x_axis.scaling.max = x_max + (0.4 * y_data_range)
            chart.y_axis.scaling.max = y_max - (0.4 * y_data_range)
            chart.x_axis.tickLblPos = "low"
            chart.x_axis.minorGridlines = ChartLines()
            chart.height = 12 
            chart.width = 22

            chart.title = None
            chart.x_axis.title = 'Frequency (Hz)'
            chart.y_axis.title = 'Magnitude (dB)'
            chart.legend.position = 'r'

            sheet.add_chart(chart, 'J2')
            workbook.save(results_file)
            
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("            Scatter Plot chart was created!              ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            
            # Create and Customize Scoring sheet
            workbook.create_sheet(result_sheet + ' Score')
            score_sheet = (result_sheet + ' Score')
            sheet = workbook[score_sheet]

            # Creating Header
            cell_ranges = ['A1:D1', 'E1:L1', 'M1:T1']
            texts = ['Info for score', 'Nimble score', 'LTspice score']

            for cell_range, text in zip(cell_ranges, texts):
                sheet.merge_cells(cell_range)
                cell = sheet[cell_range.split(':')[0]]
                cell.value = text
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Info for score table
            sheet['A2'] = 'Magnitude range'
            sheet['A3'] = y_min
            sheet['A4'] = y_max
            sheet['B2'] = 'Frequency range'
            sheet['B3'] = x_min
            sheet['B4'] = x_max
            sheet['C2'] = 'Datasheet freq'
            sheet['D2'] = 'Datasheet mag'
            # Nimble Score table
            sheet['E2'] = 'Closest match without going over index'
            sheet['F2'] = 'Below freq'
            sheet['G2'] = 'Above freq'
            sheet['H2'] = 'Below mag'
            sheet['I2'] = 'Above mag'
            sheet['J2'] = 'Linear interpolation'
            sheet['K2'] = 'Error (dB)'
            sheet['L2'] = 'Score'
            # LTspice Score table
            sheet['M2'] = 'Closest match without going over index'
            sheet['N2'] = 'Below freq'
            sheet['O2'] = 'Above freq'
            sheet['P2'] = 'Below mag'
            sheet['Q2'] = 'Above mag'
            sheet['R2'] = 'Linear interpolation'
            sheet['S2'] = 'Error (dB)'
            sheet['T2'] = 'Score'
            sheet['L2'].font = Font(bold=True)
            sheet['T2'].font = Font(bold=True)

            #Wrap cells and set width
            for col in range(1, 22):
                cell = sheet.cell(row=2, column=col)
                cell.alignment = Alignment(wrap_text=True)
                sheet.column_dimensions[get_column_letter(col)].width = 12
            
            workbook.save(results_file)
            
            #Calling the function to copy ranges within excel
            my_functions.copy_ranges_within_excel(results_file, result_sheet, score_sheet, 5, 6, 3, 4, offset_source_sheet=1, offset_target_sheet=2)

            #Calling the function to apply formulas
            my_functions.apply_formulas(results_file, score_sheet, result_sheet, x_min, x_max, y_min, y_max)
                
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("               Scoring sheet was created!                ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            
            
    def tearDown(self):
        #self.driver.quit()
        pass        

if __name__ == '__main__':
    unittest.main()
